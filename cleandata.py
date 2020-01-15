from openpyxl import load_workbook, Workbook


FILE_NAME = 'Generale.xlsx'
WORKSHEET = '_2013_11_18_tabella_per_central'
OUTPUT = 'AddressBook.xlsx'

def load(file, worksheet):
    return load_workbook(file)[worksheet]

def normalize(data):
    normalized_data = dict()


    for row in data:
        # `row[0].value != 'ID'` is just to skip the first row (the ugly way)
        if row[1].value and row[0].value != 'ID':
            # create indexes
            i = ''.join(x for x in [row[1].value, row[2].value, row[4].value, row[8].value] if x).lower()
            try:
                normalized_data[i]
            except KeyError:
                normalized_data[i] = dict()

                normalized_data[i] = {
                    'COGNOME'               : row[1].value,
                    'NOME'                  : row[2].value,
                    'non legge'             : row[3].value,
                    'NOME 2'                : row[4].value,
                    'NOTE'                  : row[5].value,
                    'AZIENDA'               : row[8].value,
                    'REPARTO'               : row[9].value,
                    'numbers'               : list()
                }

            # Check telephone number values and append to the contact's numbers list
            for n in [6,7]:
                if (row[n].value):
                    normalized_data[i]['numbers'].append(row[n].value)

    return normalized_data

def create_address_book(normalized_data, output_file):
    # create Workbook object
    wb=Workbook()

    # create worksheet
    ws = wb.active
    ws.append(['COGNOME', 'NOME', 'non legge', 'NOME 2', 'NOTE', 'AZIENDA', 'REPARTO'])
    for key in normalized_data:
        data = normalized_data[key]

        
        row = [
            data['COGNOME'], 
            data['NOME'],
            data['non legge'],
            data['NOME 2'],
            data['NOTE'],
            data['AZIENDA'],
            data['REPARTO'],
        ]


        ordered_numbers = sorted(data['numbers'], key=len)
        ws.append(row + ordered_numbers)


    wb.save(output_file)
        


def main():

    print(f"Loading data from {WORKSHEET}")
    data = load(file=FILE_NAME, worksheet=WORKSHEET)

    print("Cleaning data...")
    normalized_data = normalize(data)

    print("Creating address book...")
    create_address_book(normalized_data, OUTPUT)



if __name__ == "__main__":
    test = main()


